"""Dump every sheet of an .xlsx workbook to a readable text file.

Usage: python extract_xlsx.py <input.xlsx> <output.txt>

Called by the vault-librarian upload endpoint (deterministic server code)
so uploaded spreadsheets arrive as text the agent can Read directly.
"""
import sys

import openpyxl


def cell_str(c):
    v = c.value
    if v is None:
        return ""
    return str(v).replace("\n", " ").replace("\r", " ")


def main(src, dst):
    out = []
    for data_only in (True, False):
        wb = openpyxl.load_workbook(src, data_only=data_only, read_only=True)
        label = "VALUES" if data_only else "FORMULAS"
        if data_only:
            out.append(f"Sheets: {', '.join(wb.sheetnames)}")
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows():
                vals = [cell_str(c) for c in row]
                while vals and vals[-1] == "":
                    vals.pop()
                if vals:
                    rows.append(" | ".join(vals))
            if not rows:
                continue
            out.append(f"\n===== SHEET ({label}): {name} =====")
            out.extend(rows[:2000])
            if len(rows) > 2000:
                out.append(f"... ({len(rows) - 2000} more rows truncated)")
        wb.close()
    with open(dst, "w", encoding="utf-8") as f:
        f.write("\n".join(out))


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
