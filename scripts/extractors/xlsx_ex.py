"""Excel (.xlsx/.xlsm) extraction that preserves *meaning*, not just values.

Beyond the cell grid we capture the things a flat dump throws away and that
usually carry the real intent of an OT tracker/workbook:
  * merged cells (section headers, banners)
  * fill colours mapped to status words (RAG: red/amber/green, done, blocked)
  * formulas alongside values
  * hyperlinks, cell comments/notes
  * charts and embedded images (listed; pulled out where possible)
  * a full-page visual render of each sheet via LibreOffice
"""
from __future__ import annotations

import os

from .common import Bundle

MAX_ROWS = 4000       # per sheet safety cap for the text dump
HTML_CAP_ROWS = 800   # per sheet cap for the HTML table view


# Rough hue -> status word map for common conditional-formatting palettes.
def _status_from_rgb(rgb: str) -> str:
    if not rgb or len(rgb) < 6:
        return ""
    hexv = rgb[-6:].upper()
    try:
        r, g, bl = int(hexv[0:2], 16), int(hexv[2:4], 16), int(hexv[4:6], 16)
    except ValueError:
        return ""
    if r > 200 and g > 200 and bl > 200:
        return ""            # white/near-white = no fill meaning
    if r > 150 and g < 110 and bl < 110:
        return "RED/at-risk"
    if r > 200 and g > 150 and bl < 120:
        return "AMBER/caution"
    if g > 140 and r < 150 and bl < 150:
        return "GREEN/on-track"
    if r > 200 and g > 200 and bl < 140:
        return "YELLOW/watch"
    if r < 110 and g < 150 and bl > 180:
        return "BLUE/info"
    return f"#{hexv}"


def _fill_rgb(cell) -> str:
    try:
        fill = cell.fill
        if fill is None or fill.patternType is None:
            return ""
        col = fill.fgColor
        if col is None:
            return ""
        rgb = getattr(col, "rgb", None)
        if isinstance(rgb, str) and rgb not in ("00000000",):
            return rgb
    except Exception:
        pass
    return ""


def extract(src: str, out_dir: str, max_pages: int = 0) -> dict:
    import openpyxl

    b = Bundle(src, out_dir, "xlsx")

    # Values (computed) + formulas in one styled pass we can reuse for colours.
    wb_v = openpyxl.load_workbook(src, data_only=True)
    wb_f = openpyxl.load_workbook(src, data_only=False)

    props = wb_v.properties
    b.meta = {
        "title": props.title or "",
        "creator": props.creator or "",
        "created": str(props.created or ""),
        "modified": str(props.modified or ""),
        "sheets": list(wb_v.sheetnames),
    }
    b.add(f"# Excel workbook: {props.title or src.split('/')[-1]}")
    b.add(f"Sheets ({len(wb_v.sheetnames)}): {', '.join(wb_v.sheetnames)}")
    if props.creator:
        b.add(f"Author: {props.creator}")

    total_charts = 0
    total_images = 0
    for name in wb_v.sheetnames:
        ws_v = wb_v[name]
        ws_f = wb_f[name]
        b.heading(f"Sheet: {name}", 2)
        dims = ws_v.dimensions
        b.add(f"Used range: {dims}  ·  {ws_v.max_row} rows × {ws_v.max_column} cols")

        # Merged regions — often section headers / banners.
        merges = [str(r) for r in ws_v.merged_cells.ranges]
        if merges:
            b.add(f"Merged cells ({len(merges)}): {', '.join(merges[:60])}"
                  + (" …" if len(merges) > 60 else ""))

        # Charts & images present in the sheet.
        charts = getattr(ws_v, "_charts", []) or []
        if charts:
            total_charts += len(charts)
            titles = []
            for ch in charts:
                t = ""
                try:
                    if ch.title and ch.title.tx and ch.title.tx.rich:
                        t = "".join(
                            r.t or "" for p in ch.title.tx.rich.p for r in (p.r or [])
                        )
                except Exception:
                    pass
                titles.append(f"{type(ch).__name__.replace('Chart','')} chart"
                              + (f' "{t}"' if t else ""))
            b.add(f"Charts ({len(charts)}): {'; '.join(titles)}  "
                  "→ see the sheet render to read them.")

        images = getattr(ws_f, "_images", []) or []
        for img in images:
            total_images += 1
            try:
                data = img._data() if callable(getattr(img, "_data", None)) else None
                if data:
                    b.add_media(data, "png", desc=f"image in sheet '{name}'")
            except Exception:
                pass
        if images:
            b.add(f"Embedded images: {len(images)}")

        # Cell grid: values, formulas where present, status from fill colour.
        b.add("", "Data (value cells; `=formula` and [status] shown where set):")
        rows_out = 0
        status_notes: list[str] = []
        for r_v, r_f in zip(ws_v.iter_rows(), ws_f.iter_rows()):
            cells = []
            any_val = False
            for c_v, c_f in zip(r_v, r_f):
                val = c_v.value
                txt = "" if val is None else str(val).replace("\n", " ").replace("|", "/")
                # formula
                if isinstance(c_f.value, str) and c_f.value.startswith("="):
                    txt = f"{txt} ({c_f.value})" if txt else c_f.value
                # status from fill colour
                rgb = _fill_rgb(c_f)
                if rgb:
                    st = _status_from_rgb(rgb)
                    if st:
                        txt = f"{txt} [{st}]" if txt else f"[{st}]"
                        if c_v.value is not None:
                            status_notes.append(f"{c_v.coordinate} {c_v.value!r} → {st}")
                # comments
                if c_f.comment is not None and c_f.comment.text:
                    txt += f" «note: {c_f.comment.text.strip()[:120]}»"
                # hyperlink
                if c_v.hyperlink is not None and c_v.hyperlink.target:
                    txt += f" <{c_v.hyperlink.target}>"
                if txt:
                    any_val = True
                cells.append(txt)
            while cells and cells[-1] == "":
                cells.pop()
            if any_val:
                b.add("| " + " | ".join(cells) + " |")
                rows_out += 1
            if rows_out >= MAX_ROWS:
                b.add(f"… (row cap {MAX_ROWS} reached for this sheet)")
                break

        if status_notes:
            b.add("", f"Status-coloured cells ({len(status_notes)}): "
                  + "; ".join(status_notes[:40]) + (" …" if len(status_notes) > 40 else ""))

    b.meta["charts"] = total_charts
    b.meta["embeddedImages"] = total_images

    # Rich HTML tables (all columns, fill colours, merged cells) — far more usable
    # than paginating a wide sheet into narrow PDF pages (LibreOffice's default).
    try:
        b.add_html(_build_html(wb_v, props.title or os.path.basename(src)))
    except Exception as e:
        b.warn(f"html view: {e}")

    wb_v.close()
    wb_f.close()
    return b.result()


def _esc(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


_HTML_HEAD = """<!doctype html><meta charset="utf-8">
<style>
:root{color-scheme:light dark}
body{margin:0;font:12px/1.4 -apple-system,Segoe UI,Roboto,sans-serif;background:#fff;color:#141821}
@media(prefers-color-scheme:dark){body{background:#0f1319;color:#e6eaf0}}
h2{font:600 13px/1 sans-serif;margin:16px 12px 6px;position:sticky;left:0}
.wrap{overflow:auto;margin:0 10px 18px;border:1px solid #8884;border-radius:6px;max-height:none}
table{border-collapse:collapse;white-space:nowrap}
td{border:1px solid #8883;padding:2px 7px;max-width:360px;overflow:hidden;text-overflow:ellipsis}
tr:first-child td{font-weight:600;background:#8881}
.more{color:#8899;font-style:italic}
</style>"""


def _build_html(wb, title):
    parts = [_HTML_HEAD, f"<title>{_esc(title)}</title>"]
    for name in wb.sheetnames:
        ws = wb[name]
        merged_tl, covered = {}, set()
        for rng in ws.merged_cells.ranges:
            merged_tl[(rng.min_row, rng.min_col)] = (rng.max_row - rng.min_row + 1, rng.max_col - rng.min_col + 1)
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if (r, c) != (rng.min_row, rng.min_col):
                        covered.add((r, c))
        parts.append(f"<h2>{_esc(name)}</h2><div class='wrap'><table>")
        rown = 0
        for row in ws.iter_rows():
            rown += 1
            if rown > HTML_CAP_ROWS:
                parts.append(f"<tr><td class='more'>… {ws.max_row - HTML_CAP_ROWS} more rows — open the original file for the full sheet</td></tr>")
                break
            tds = []
            for cell in row:
                rc = (cell.row, cell.column)
                if rc in covered:
                    continue
                txt = "" if cell.value is None else _esc(cell.value)
                attrs = ""
                span = merged_tl.get(rc)
                if span:
                    if span[0] > 1:
                        attrs += f" rowspan={span[0]}"
                    if span[1] > 1:
                        attrs += f" colspan={span[1]}"
                rgb = _fill_rgb(cell)
                if rgb:
                    attrs += f" style=\"background:#{rgb[-6:]};color:#141821\""
                tds.append(f"<td{attrs}>{txt}</td>")
            parts.append("<tr>" + "".join(tds) + "</tr>")
        parts.append("</table></div>")
    return "".join(parts)
